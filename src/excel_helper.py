import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from datetime import datetime, timedelta
import pytz

# This map handles the names and units for your fixed 15-item list.
# This ensures that even with 0 orders, the row is correctly labeled.
MASTER_SUPPLY_MAP = {
    "e07f0dd7-bd01-42d2-8d9b-6c80813a6d81": {"name": 'PAN 1/3 SIZE 4"', "unit": "Each"},
    "2d31b938-3d19-48bc-8480-510bc5df6672": {"name": 'LID 1/3 SIZE', "unit": "Each"},
    "cb7ffc75-819c-41aa-8f88-17bdaf07fec2": {"name": 'PAN HALF SIZE 2"', "unit": "Each"},
    "4759978e-88fb-49a8-9e46-5accc4e4a7b3": {"name": 'PAN HALF SIZE 4"', "unit": "Each"},
    "ec24d5fc-e7ed-4a5e-a522-74ed46cb6c4e": {"name": 'LID HALF SIZE', "unit": "Each"},
    "9f48f5d1-0cf5-4d4d-bc60-dcee366bfd29": {"name": 'PAN FULL SIZE 6"', "unit": "Each"},
    "82487c9e-069d-429d-bfd6-0266e7322708": {"name": 'LID FULL SIZE', "unit": "Each"},
    "80ac6d73-e615-4faa-b54b-17631a2395e3": {"name": 'SERVING FORK BLK', "unit": "Each"},
    "2138eec0-233b-49cf-9062-65eecec4b6f3": {"name": 'SERVING SPOON BLK', "unit": "Each"},
    "40e2d16f-4bdb-421b-8f05-8ed036526aee": {"name": 'SERVING TONG BLK', "unit": "Each"},
    "c31bf9a5-e965-4378-9f4f-00772c43bd64": {"name": 'BOWL 160OZ', "unit": "Each"},
    "ef1f28a2-3dd3-4b42-b9e0-1293f1954a0c": {"name": 'LID BOWL 160OZ', "unit": "Each"},
    "67ac8a0a-1838-41b5-bb94-875b1d71ede3": {"name": 'LID BOWL 64OZ', "unit": "Each"},
    "b296f9ef-d42c-42e3-a342-5694a1cc23ff": {"name": 'BOWL 64OZ', "unit": "Each"},
    "d299c5d1-bc1f-43c6-83ac-cf22851bd201": {"name": 'PLATE 10" BLK', "unit": "Each"},
}

PAR_VALUES = {
    'PAN 1/3 SIZE 4"': 20, 'LID 1/3 SIZE': 20, 'PAN HALF SIZE 2"': 6,
    'PAN HALF SIZE 4"': 50, 'LID HALF SIZE': 50, 'PAN FULL SIZE 6"': 15,
    'LID FULL SIZE': 10, 'SERVING FORK BLK': 48, 'SERVING SPOON BLK': 48,
    'SERVING TONG BLK': 48, 'BOWL 160OZ': 12, 'LID BOWL 160OZ': 12,
    'LID BOWL 64OZ': 12, 'BOWL 64OZ': 12, 'PLATE 10" BLK': 2,
    # Additional Par items (if appearing in DB)
    'CHAFER RACK': 13, 'STERNO': 25, 'COFFEE BOX 160OZ': 7,
    'COFFEE BOX 3 GALLON': 5, 'CUP COFFEE 12OZ': 6, 'TABLE CLOTH ROLL': 2,
}

FIXED_GUID_ORDER = list(MASTER_SUPPLY_MAP.keys())

THIN = Side(border_style='thin', color='000000')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _font(bold=False, size=8):
    return Font(name='Calibri', bold=bold, size=size)


def _align(horizontal='left', vertical='center', wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _write(ws, row, col, value, font=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    return cell


def get_excel_download_buffer(df, order_df, sheet_name="AS CATERING - NEW", location_name="Burke",
                              report_date=datetime(2026, 1, 1, 0, 0, 0), route=""):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # --- HEADERS ---
    ws.oddHeader.left.text = f'&"Times New Roman,Bold"&16{route.upper()}'
    ws.oddHeader.right.text = f'&"Times New Roman,Bold"&16{(report_date - timedelta(days=1)).strftime('%a').upper()} for {report_date.strftime('%a').upper()}\n&"Times New Roman,Regular"&10 {report_date.strftime('%m/%d/%Y')}'
    ws.oddHeader.center.text = f'&"Times New Roman,Bold"&16{location_name} - CATERING SUPPLIES'

    bold9, reg8, bold8 = _font(bold=True, size=9), _font(bold=False, size=8), _font(bold=True, size=8)

    # Row 1: Name / Date
    _write(ws, 1, 1, 'NAME', font=bold9, alignment=_align('right'))
    ws.merge_cells('B1:G1')
    for col in range(2, 8): ws.cell(1, col).border = Border(bottom=THIN)

    _write(ws, 1, 10, 'DATE', font=bold9, alignment=_align('right'))
    ny_tz = pytz.timezone('America/New_York')
    ws.merge_cells('K1:N1')
    _write(ws, 1, 11, datetime.now(ny_tz).strftime('%m/%d/%Y %I:%M %p'), font=reg8, alignment=_align('left'))
    for col in range(11, 15): ws.cell(1, col).border = Border(bottom=THIN)

    # Row 2: MOD
    _write(ws, 2, 1, 'MOD', font=bold9, alignment=_align('right'))
    ws.merge_cells('B2:G2')
    for col in range(2, 8): ws.cell(2, col).border = Border(bottom=THIN)
    _write(ws, 2, 10, 'TIME OF INV.', font=bold9, alignment=_align('right'))
    ws.merge_cells('K2:N2')
    for col in range(11, 15): ws.cell(2, col).border = Border(bottom=THIN)

    ws.row_dimensions[3].height = 6

    # --- TABLE HEADER ---
    headers = ['SUPPLIES', 'UNIT', 'INV', 'PAR', 'ORD', '√', '√', 'B/O']
    for col_idx, h in enumerate(headers, start=1):
        _write(ws, 4, col_idx, h, font=bold8, alignment=_align('center'), border=THIN_BORDER)

    # --- DATA MAPPING ---
    incoming_data = {str(row.get('supply_id')): row for _, row in df.iterrows()} if not df.empty else {}

    curr_row = 5
    processed_ids = set()

    # 1. FIXED SEQUENCE LOOP (Ensures all 15 items show up even if 0 ordered)
    for s_id in FIXED_GUID_ORDER:
        item = incoming_data.get(s_id)
        master_info = MASTER_SUPPLY_MAP.get(s_id)

        name = str(item['Supply Item']).upper() if item is not None else master_info['name'].upper()
        unit = str(item['Type']) if item is not None else master_info['unit']
        qty = item['Total Qty'] if item is not None else 0
        ###hold over incase we want to add par values later###
        #par = PAR_VALUES.get(name, "")

        _write(ws, curr_row, 1, name, font=reg8, border=THIN_BORDER)
        _write(ws, curr_row, 2, unit, font=reg8, alignment=_align('center'), border=THIN_BORDER)
        _write(ws, curr_row, 3, "", border=THIN_BORDER)
        _write(ws, curr_row, 4, "", font=reg8, alignment=_align('center'), border=THIN_BORDER)
        _write(ws, curr_row, 5, qty if qty > 0 else "", font=bold8, alignment=_align('center'), border=THIN_BORDER)
        for c in [6, 7, 8]: _write(ws, curr_row, c, "", border=THIN_BORDER)

        processed_ids.add(s_id)
        curr_row += 1

    # 2. EXTRA ITEMS (Catch-all for anything else from the database)
    for s_id, row in incoming_data.items():
        if s_id not in processed_ids:
            _write(ws, curr_row, 1, str(row['Supply Item']).upper(), font=reg8, border=THIN_BORDER)
            _write(ws, curr_row, 2, str(row['Type']), font=reg8, alignment=_align('center'), border=THIN_BORDER)
            _write(ws, curr_row, 5, row['Total Qty'], font=bold8, alignment=_align('center'), border=THIN_BORDER)
            for c in [3, 4, 6, 7, 8]: _write(ws, curr_row, c, "", border=THIN_BORDER)
            curr_row += 1

    # --- OPERATIONAL SECTION ---
    curr_row += 1
    ops_headers = ['OPERATIONAL', 'UNIT', 'INV', 'PAR', 'ORD', '√', '√', 'B/O']
    for col_idx, header in enumerate(ops_headers, start=1):
        align = _align('left') if col_idx == 1 else _align('center')
        _write(ws, curr_row, col_idx, header, font=bold8, alignment=align, border=THIN_BORDER)
    curr_row += 1

    for item_name in ['CAMBROS', 'COFFEE CAMBROS', 'VAN']:
        _write(ws, curr_row, 1, item_name, font=reg8, alignment=_align('left'), border=THIN_BORDER)
        _write(ws, curr_row, 2, 'EA' if item_name != 'VAN' else '', font=reg8, alignment=_align('center'),
               border=THIN_BORDER)
        for col_idx in range(3, 9): _write(ws, curr_row, col_idx, '', border=THIN_BORDER)
        curr_row += 1

    # --- INSTRUCTIONS ---
    curr_row += 2
    _write(ws, curr_row, 1, 'INSTRUCTIONS:', font=bold8)
    for i, text in enumerate(['1.- TO BE SENT ON TUESDAY FOR NEXT DELIVERY DAY',
                              '2.- REGARDLESS OF ORDER, YOU MUST STILL EMAIL INVENTORY',
                              '3 - SEND TO COMMISSARY@ANITASCORP.COM']):
        _write(ws, curr_row + i + 1, 1, text, font=reg8)

    # --- ORDER SUMMARY ---
    curr_row += 5
    if order_df is not None and not order_df.empty:
        _write(ws, curr_row, 1, 'DAILY ORDER SUMMARY:', font=bold8)
        curr_row += 1
        for _, group in order_df.groupby('order_guid', sort=False):
            meta = group.iloc[0]
            d_opt = str(meta.get('dining_option', 'N/A')).upper()
            summary = f"({meta['local_time'].strftime('%I:%M %p')}) #{meta['order_number']} - {str(meta['customer_name']).upper()} - ${meta['order_total']:,.2f} [{d_opt}]"
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
            _write(ws, curr_row, 1, summary, font=reg8)
            for col in range(1, 9): ws.cell(curr_row, col).border = Border(bottom=Side(style='hair'))
            curr_row += 1

    # --- PAGE SETUP ---
    widths = {'A': 20.0, 'B': 7.29, 'C': 4.57, 'D': 3.71, 'E': 4.57, 'F': 3.86, 'G': 3.86, 'H': 3.86, 'I': 1.0,
              'J': 14.14, 'K': 5.29, 'L': 4.57, 'M': 3.71, 'N': 4.57}
    for col, w in widths.items(): ws.column_dimensions[col].width = w
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=1, bottom=0.5)

    # --- INDIVIDUAL PACKING SLIPS ---
    if order_df is not None and not order_df.empty:
        for _, group in order_df.groupby('order_guid', sort=False):
            meta = group.iloc[0]
            ws_order = wb.create_sheet(title=f"Order {meta['order_number']}")
            _write(ws_order, 1, 1, "PACKING SLIP", font=_font(bold=True, size=14))
            _write(ws_order, 2, 1, f"CUSTOMER: {str(meta['customer_name']).upper()}")
            _write(ws_order, 3, 1, f"ORDER #: {meta['order_number']}")
            _write(ws_order, 4, 1, f"DATE: {meta['local_time'].strftime('%b %d,%Y')}")
            _write(ws_order, 5, 1, f"TIME: {meta['local_time'].strftime('%I:%M %p')}")
            _write(ws_order, 7, 1, "REQUIRED GEAR", font=bold8)
            _write(ws_order, 8, 1, "ITEM", font=bold8, border=THIN_BORDER)
            _write(ws_order, 8, 2, "QTY", font=bold8, border=THIN_BORDER)
            detail_row = 9
            for _, s_row in group.iterrows():
                _write(ws_order, detail_row, 1, s_row['supply_name'], font=reg8, border=THIN_BORDER)
                _write(ws_order, detail_row, 2, s_row['units_needed'], font=bold8, border=THIN_BORDER)
                detail_row += 1
            ws_order.column_dimensions['A'].width = 30

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()